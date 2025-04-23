import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from collections import defaultdict

def process_excel():
    # 读取Excel文件
    df = pd.read_excel('venv/file/云赛数据.xls')
    
    # 筛选铝型材数据
    df = df[df['物料名称'] == '铝型材']
    
    # 处理描述字段，按逗号拆分
    def split_description(desc):
        if isinstance(desc, str):
            return [item.strip() for item in desc.split(',')]
        return []
    
    # 创建结果DataFrame
    result_data = []
    
    # 获取所有描述项的最大数量
    max_desc_items = max(len(split_description(desc)) for desc in df['描述'].unique())
    
    # 处理每一行数据
    for _, row in df.iterrows():
        desc_items = split_description(row['描述'])
        
        # 创建基础结果行
        result_row = {
            '品牌': row['品牌'],
            '供应商': row['供应商'],
            '描述': row['描述']
        }
        
        # 定义描述项的列名映射
        desc_column_names = ['属性码1', '属性码2', '长度', '加强筋', '溢胶槽', '单位', '开孔配置', '制造工艺', '边框槽', 'A面外观', 'C面', '文件编号', '膜厚']
        
        # 添加拆分后的描述列
        for i in range(max_desc_items):
            col_name = desc_column_names[i] if i < len(desc_column_names) else f'描述项_{i+1}'
            result_row[col_name] = desc_items[i] if i < len(desc_items) else ''
        
        # 添加日期和价格
        date = pd.to_datetime(row['采购日期']).strftime('%Y-%m-%d')
        result_row[date] = row['含税单价']
        
        result_data.append(result_row)
    
    # 创建结果DataFrame
    result_df = pd.DataFrame(result_data)
    
    # 对日期列进行排序和月份合并处理
    date_columns = [col for col in result_df.columns if col not in ['品牌', '供应商', '描述'] + 
                   [desc_column_names[i] if i < len(desc_column_names) else f'描述项_{i+1}' for i in range(max_desc_items)]]
    date_columns = sorted(date_columns)
    
    # 创建月份映射
    month_data = defaultdict(list)
    for date_col in date_columns:
        month = pd.to_datetime(date_col).strftime('%Y-%m')
        month_data[month].append(date_col)
    
    # 处理月份合并
    merged_columns = []
    for month, dates in month_data.items():
        all_prices = result_df[dates].values.flatten()
        all_prices = all_prices[~pd.isna(all_prices)]
        
        if all(abs(price - 0.00001) < 1e-10 for price in all_prices):
            # 所有价格都是0.00001，使用月份
            merged_df = result_df[dates].copy()
            merged_df.columns = [month] * len(dates)
            result_df[month] = merged_df.iloc[:, 0]
            for date in dates[1:]:
                result_df.drop(columns=[date], inplace=True)
            merged_columns.append(month)
        else:
            # 保持原始日期
            merged_columns.extend(dates)
    
    # 重新排序列
    result_df = result_df.reindex(columns=['品牌', '供应商', '描述'] + 
                                 [desc_column_names[i] if i < len(desc_column_names) else f'描述项_{i+1}' for i in range(max_desc_items)] + 
                                 merged_columns)
    
    # 保存到新的Excel文件
    output_file = 'venv/file/铝型材分类结果.xlsx'
    result_df.to_excel(output_file, index=False)
    
    # 加载工作簿进行单元格合并
    wb = load_workbook(output_file)
    ws = wb.active
    
    # 获取需要合并的列
    merge_columns = ['品牌', '供应商', '描述']
    merge_columns.extend([desc_column_names[i] if i < len(desc_column_names) else f'描述项_{i+1}' for i in range(max_desc_items)])
    
    # 对每个需要合并的列进行处理
    for col_idx, col_name in enumerate(merge_columns, start=1):
        start_row = 2  # 从第二行开始（跳过标题行）
        current_value = ws.cell(row=start_row, column=col_idx).value
        
        for row in range(3, ws.max_row + 1):
            if ws.cell(row=row, column=col_idx).value == current_value:
                continue
            
            if row - 1 > start_row:
                ws.merge_cells(start_row=start_row, start_column=col_idx,
                              end_row=row-1, end_column=col_idx)
            
            start_row = row
            current_value = ws.cell(row=row, column=col_idx).value
        
        # 处理最后一组相同的值
        if ws.max_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=col_idx,
                          end_row=ws.max_row, end_column=col_idx)
    
    # 设置单元格样式和冻结窗格
    ws.freeze_panes = 'A2'  # 冻结第一行
    date_start_col = len(merge_columns) + 1
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 处理日期列的样式
            if cell.column >= date_start_col:
                if cell.row > 1 and cell.value is not None:  # 跳过标题行
                    try:
                        value = float(cell.value)
                        # 只有当月份被合并且值为0.00001时才标红
                        col_letter = get_column_letter(cell.column)
                        col_name = ws[f'{col_letter}1'].value
                        if col_name in merged_columns and abs(value - 0.00001) < 1e-10:
                            cell.font = Font(color='FF0000')
                    except (ValueError, TypeError):
                        pass
    
    # 保存修改后的文件
    wb.save(output_file)

if __name__ == '__main__':
    process_excel()
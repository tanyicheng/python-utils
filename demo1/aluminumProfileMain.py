import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from collections import defaultdict

def process_excel():
    input_file = 'venv/file/安赛数据.xls'
    df = pd.read_excel(input_file)
    df = df[df['物料名称'] == '铝型材']

    def split_description(desc):
        if isinstance(desc, str):
            return [item.strip() for item in desc.split(',')]
        return []

    desc_column_names = ['颜色', '宽度', '长度', '加强筋', '溢胶槽', '单位', '开孔配置', '制造工艺', '边框槽', 'A面外观', 'C面', '文件编号', '膜厚']
    max_desc_items = max(len(split_description(desc)) for desc in df['描述'].unique())

    # 分组：品牌、供应商、描述、所有描述项
    group_cols = ['品牌', '供应商', '描述'] + [desc_column_names[i] if i < len(desc_column_names) else f'描述项_{i+1}' for i in range(max_desc_items)]
    result_rows = []
    price_col_info = []  # (列名, 是否多价, 日期/月份, 是否标红)
    all_dates = sorted(df['采购日期'].apply(lambda x: pd.to_datetime(x).strftime('%Y-%m-%d')).unique())
    all_months = sorted(set([d[:7] for d in all_dates]))

    # 预处理所有描述项
    df_desc = df.copy()
    for i in range(max_desc_items):
        df_desc[f'desc_{i}'] = df_desc['描述'].apply(lambda x: split_description(x)[i] if i < len(split_description(x)) else '')

    # 分组聚合
    grouped = df_desc.groupby(['品牌', '供应商', '描述'] + [f'desc_{i}' for i in range(max_desc_items)])
    for group_keys, group_df in grouped:
        base_row = {
            '品牌': group_keys[0],
            '供应商': group_keys[1],
            '描述': group_keys[2]
        }
        for i in range(max_desc_items):
            col_name = desc_column_names[i] if i < len(desc_column_names) else f'描述项_{i+1}'
            base_row[col_name] = group_keys[3+i]
        
        # 按月分组
        month_group = group_df.groupby(group_df['采购日期'].apply(lambda x: pd.to_datetime(x).strftime('%Y-%m')))
        price_cells = {}
        
        for month, month_df in month_group:
            # 获取该月份的所有不同价格
            month_prices = month_df['含税单价'].unique()
            
            if len(month_prices) == 1:
                # 如果该月只有一个价格，使用月份作为列名
                # 获取第一条记录的日期
                first_date = pd.to_datetime(month_df['采购日期'].iloc[0]).strftime('%Y-%m-%d')
                price_cells[month] = {'value': month_prices[0], 'red': False}
                price_col_info.append((month, False, first_date, False))
            else:
                # 如果该月有多个不同价格，按日期分别显示
                for _, row in month_df.iterrows():
                    date = pd.to_datetime(row['采购日期']).strftime('%Y-%m-%d')
                    price_cells[date] = {'value': row['含税单价'], 'red': False}
                    price_col_info.append((date, True, date, False))
        
        row = base_row.copy()
        row.update({k: v['value'] for k, v in price_cells.items()})
        result_rows.append(row)

    # 确定所有需要的列顺序
    col_set = set()
    for col, is_multi, label, is_red in price_col_info:
        col_set.add(col)
    
    # 按时间排序，月份在前，日期在后
    all_cols = sorted(col_set)
    month_cols = [c for c in all_cols if len(c)==7]  # YYYY-MM 格式
    date_cols = [c for c in all_cols if len(c)==10]  # YYYY-MM-DD 格式
    
    # 重新组织列顺序：描述相关列 -> 日期列 -> 品牌供应商
    desc_related_cols = ['描述'] + [desc_column_names[i] if i < len(desc_column_names) else f'描述项_{i+1}' for i in range(max_desc_items)]
    brand_supplier_cols = ['品牌', '供应商']
    
    # 新的列顺序：描述相关列 + 品牌供应商 + 月份 + 日期
    final_cols = desc_related_cols + brand_supplier_cols + month_cols + date_cols
    result_df = pd.DataFrame(result_rows)
    result_df = result_df.reindex(columns=final_cols)

    # 修改输出文件名：使用输入文件名 + 分析结果
    input_filename = input_file.split('/')[-1].split('.')[0]  # 获取输入文件名(不含扩展名)
    output_file = f'venv/file/{input_filename}铝型材数据分析.xlsx'
    result_df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    # 修改月份列的标题，只显示日期
    month_dates = {}  # 用于存储每个月份列的第一个日期
    
    # 先遍历所有数据行，找到每个月份列的第一个有效日期
    for col_idx, col_name in enumerate(final_cols, start=1):
        if len(col_name) == 7:  # 是月份列
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value is not None:
                    # 找到第一个有值的单元格，记录对应的日期
                    month_dates[col_name] = price_col_info[next(i for i, (col, _, date, _) 
                                                              in enumerate(price_col_info) 
                                                              if col == col_name)][2]
                    break
    
    # 更新月份列标题，只显示日期
    for col_idx, col_name in enumerate(final_cols, start=1):
        if len(col_name) == 7 and col_name in month_dates:  # 是月份列且有对应日期
            cell = ws.cell(row=1, column=col_idx)
            cell.value = month_dates[col_name]  # 直接使用日期作为标题

    # 合并规格、长度、颜色相同的行
    merge_cols = {
        '规格': None,
        '长度': None,
        '颜色': None
    }
    
    # 找到这些列的索引
    for col_idx, col_name in enumerate(final_cols, start=1):  # 使用 final_cols 而不是 group_cols
        if col_name in merge_cols:
            merge_cols[col_name] = col_idx

    # 合并相同数据的单元格
    for col_name, col_idx in merge_cols.items():
        if col_idx is None:
            continue
        start_row = 2
        current_value = ws.cell(row=start_row, column=col_idx).value
        for row in range(3, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value == current_value:
                continue
            if row - 1 > start_row:
                ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=row-1, end_column=col_idx)
            start_row = row
            current_value = cell_value
        # 处理最后一组相同的值
        if ws.max_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=ws.max_row, end_column=col_idx)

    # 修改斑马线效果为浅蓝色
    light_blue_fill = PatternFill(start_color='DCE6F2', end_color='DCE6F2', fill_type='solid')  # 使用浅蓝色
    for row in range(2, ws.max_row + 1):  # 从第二行开始（跳过标题行）
        if row % 2 == 0:  # 偶数行添加浅蓝色背景
            for cell in ws[row]:
                cell.fill = light_blue_fill

    # 定义白色边框样式
    white_side = Side(style='thin', color='FFFFFF')
    white_border = Border(
        left=white_side,
        right=white_side,
        top=white_side,
        bottom=white_side
    )

    # 设置标题行背景色为深蓝色，添加白色边框
    header_blue_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_blue_fill
        cell.font = header_font
        cell.border = white_border

    # 为所有数据单元格添加白色边框
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = white_border

    # 设置其他样式
    ws.freeze_panes = 'A2'
    date_start_col = len(group_cols) + 1

    # 自动调整列宽和内容居中
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width
    
    # 居中对齐
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(output_file)

if __name__ == '__main__':
    process_excel()